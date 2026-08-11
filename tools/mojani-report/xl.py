from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,PatternFill,Border,Side
from openpyxl.utils import get_column_letter
FT=3.28084; SQFT=10.7639; G=101.17
H=Font(bold=True,color='FFFFFF'); HF=PatternFill('solid',fgColor='1A231C')
T=Font(bold=True); TF=PatternFill('solid',fgColor='EDEFE8')
thin=Side(style='thin',color='C7CEC0'); BD=Border(top=thin,bottom=thin,left=thin,right=thin)
wb=Workbook()

def sheet(ws,title,head,rows,widths,foot=None):
    ws.append([title]); ws['A1'].font=Font(bold=True,size=13); ws.append([])
    ws.append(head)
    for c in range(1,len(head)+1):
        cell=ws.cell(row=3,column=c); cell.font=H; cell.fill=HF; cell.border=BD
        cell.alignment=Alignment(horizontal='center',wrap_text=True)
    for r in rows:
        ws.append(r)
    if foot:
        ws.append(foot)
        for c in range(1,len(head)+1):
            cell=ws.cell(row=ws.max_row,column=c); cell.font=T; cell.fill=TF
    for row in ws.iter_rows(min_row=4,max_row=ws.max_row,max_col=len(head)):
        for cell in row:
            cell.border=BD
            if isinstance(cell.value,(int,float)): cell.alignment=Alignment(horizontal='right')
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A4'

ws=wb.active; ws.title='सारांश'
def R2(a,b,ar,pos,nm):
    return [nm[0],nm[1],round(a*FT,2),round(b*FT,2),round(ar*SQFT,2),round(a,2),round(b,2),ar,round(ar/G,3),pos]
rows=[R2(28.79,5.70,164.06,'दक्षिणेकडील पट्टा',['2457/1','श्री. नेमिनाथ आण्णाप्पा वसवाडे']),
      R2(25.91,2.40,62.30,'मधला पट्टा',['2457/2','श्री. देवाप्पा आण्णाप्पा वसवाडे']),
      R2(25.50,2.44,62.30,'उत्तरेकडील पट्टा',['2457/3','श्री. आदिनाथ आण्णाप्पा वसवाडे']),
      R2(5.35,1.97,10.54,'पश्चिम टोकाचा सामाईक तुकडा',['2457/4','तिघे सामाईक (नेमिनाथ, देवाप्पा, आदिनाथ)'])]
sheet(ws,'मौजे हुपरी, ता. हातकणंगले, जि. कोल्हापूर — सि.स.नं. 2457 पोटहिस्सा क्षेत्र तक्ता',
 ['सि.स.नं.','धारकाचे नाव','सरासरी लांबी (फूट)','सरासरी रुंदी (फूट)','क्षेत्र (चौ.फूट)',
  'सरासरी लांबी (मी.)','सरासरी रुंदी (मी.)','क्षेत्र (चौ.मी.)','गुंठे','स्थान'],
 rows,[12,36,17,17,15,17,17,14,10,28],
 ['एकूण','—','','',3220.56,'','',299.20,2.957,'मुख्य सि.स.नं. 2457'])

ws2=wb.create_sheet('मुख्य प्लॉट बाजू')
_m=[('A–B','उत्तर बाजू — मुख्य',25.28,'सि.स.नं. 2453/5 व 2460'),
 ('B–C','उत्तर बाजू — पश्चिम टोक',1.93,'सि.स.नं. 2460'),
 ('C–D','पश्चिम बाजू — उत्तर भाग',5.35,'सि.स.नं. 2461'),
 ('D–E','पश्चिम बाजू — दक्षिण भाग',5.74,'सि.स.नं. 2463'),
 ('E–F','दक्षिण बाजू — मुख्य',29.37,'सि.स.नं. 2496/1, 2496/2, 2496/3'),
 ('F–G','पूर्व बाजू — दक्षिण भाग',5.78,'लागू रस्ता'),
 ('G–H','पूर्व बाजू — मधला भाग',2.36,'लागू रस्ता'),
 ('H–A','पूर्व बाजू — उत्तर भाग',2.10,'लागू रस्ता')]
main=[[a,b,round(m*FT,2),m,e] for a,b,m,e in _m]
sheet(ws2,'मुख्य मिळकत सि.स.नं. 2457 — बाजूनिहाय मापे',
 ['बाजू','दिशा','लांबी (फूट)','लांबी (मी.)','लगत मिळकत'],main,[10,28,14,14,34],
 ['एकूण परिमिती','—',255.61,77.91,'क्षेत्र 3,220.56 चौ.फूट (299.20 चौ.मी.)'])

ws3=wb.create_sheet('पोटहिस्सा बाजू')
sub=[('2457/1','D–G','उत्तर — अंतर्गत हद्द',28.21,'सि.स.नं. 2457/4 व 2457/2'),
 ('2457/1','G–F','पूर्व',5.78,'लागू रस्ता'),
 ('2457/1','F–E','दक्षिण',29.37,'सि.स.नं. 2496/1, 2496/2, 2496/3'),
 ('2457/1','E–D','पश्चिम',5.74,'सि.स.नं. 2463'),
 ('2457/2','P1–H','उत्तर — अंतर्गत हद्द',25.71,'सि.स.नं. 2457/3'),
 ('2457/2','H–G','पूर्व',2.36,'लागू रस्ता'),
 ('2457/2','G–P2','दक्षिण — अंतर्गत हद्द',26.11,'सि.स.नं. 2457/1'),
 ('2457/2','P2–P1','पश्चिम — अंतर्गत हद्द',2.46,'सि.स.नं. 2457/4'),
 ('2457/3','B–A','उत्तर',25.28,'सि.स.नं. 2460 व 2453/5'),
 ('2457/3','A–H','पूर्व',2.10,'लागू रस्ता'),
 ('2457/3','H–P1','दक्षिण — अंतर्गत हद्द',25.71,'सि.स.नं. 2457/2'),
 ('2457/3','P1–B','पश्चिम — अंतर्गत हद्द',2.81,'सि.स.नं. 2457/4'),
 ('2457/4','C–B','उत्तर',1.93,'सि.स.नं. 2460'),
 ('2457/4','B–P1','पूर्व — अंतर्गत हद्द',2.81,'सि.स.नं. 2457/3'),
 ('2457/4','P1–P2','पूर्व — अंतर्गत हद्द',2.46,'सि.स.नं. 2457/2'),
 ('2457/4','P2–D','दक्षिण — अंतर्गत हद्द',2.10,'सि.स.नं. 2457/1'),
 ('2457/4','D–C','पश्चिम',5.35,'सि.स.नं. 2461')]
sheet(ws3,'प्रत्येक पोटहिस्स्याची स्वतंत्र बाजूनिहाय मापे',
 ['सि.स.नं.','बाजू','दिशा / प्रकार','लांबी (फूट)','लांबी (मी.)','लगत'],
 [[a,b,c,round(d*FT,2),d,e] for a,b,c,d,e in sub],[12,10,26,14,14,34])

# --- मूल्यांकन (rates.json असल्यासच) ---
import os, json
if os.path.exists('rates.json'):
    _cfg=json.load(open('rates.json',encoding='utf-8'))
    if _cfg.get('asr_rate_per_sqm') or _cfg.get('market_rate_per_sqft'):
        from valuation import compute
        _rows,_tot=compute(_cfg)
        wsv=wb.create_sheet('मूल्यांकन')
        _hd=['सि.स.नं.','धारक','क्षेत्र (चौ.फूट)','क्षेत्र (चौ.मी.)','घटक']
        if _cfg.get('asr_rate_per_sqm'): _hd.append('शासकीय मूल्य ₹')
        if _cfg.get('market_rate_per_sqft'): _hd.append('बाजार मूल्य ₹')
        _data=[]
        for r in _rows:
            row=[r['no'],r['holder'],round(r['sqft'],2),round(r['sqm'],2),r['factor']]
            if _cfg.get('asr_rate_per_sqm'): row.append(round(r['asr_value']))
            if _cfg.get('market_rate_per_sqft'): row.append(round(r['mkt_value']))
            _data.append(row)
        _ft=['एकूण','—',round(_tot['sqft'],2),round(_tot['sqm'],2),'']
        if _cfg.get('asr_rate_per_sqm'): _ft.append(round(_tot['asr']))
        if _cfg.get('market_rate_per_sqft'): _ft.append(round(_tot['mkt']))
        _title='मूल्यांकन (सूचक) — शासकीय दर ₹%s/चौ.मी.'%_cfg.get('asr_rate_per_sqm','—')
        sheet(wsv,_title,_hd,_data,[12,36,16,15,9,18,18],_ft)
        wsv.append([])
        wsv.append(['स्रोत',_cfg.get('asr_source','')])
        wsv.append(['अस्वीकरण','हे मूल्य सूचक (indicative) आहे. मुद्रांक शुल्कासाठीचे अधिकृत बाजारमूल्य '
                    'नोंदणी व मुद्रांक विभागाच्या दर तक्त्यानुसार व शासनमान्य मूल्यांकनकर्त्याकडूनच ठरते.'])
        for _r in wsv.iter_rows(min_row=wsv.max_row-1,max_row=wsv.max_row):
            _r[1].alignment=Alignment(wrap_text=True,vertical='top')

ws4=wb.create_sheet('नोंदी')
for r in [['मोजणी तपशील'],[],
 ['मौजे','हुपरी'],['तालुका','हातकणंगले'],['जिल्हा','कोल्हापूर'],
 ['मोजणी रजिस्टर नं.','090729/2025 (नियमित पोटहिस्सा)'],['मोजणी दिनांक','13/03/2026'],
 ['प्रमाण','1:500'],['प्रत','अ प्रत'],
 ['अर्जदार','श्री. नेमिनाथ आण्णाप्पा वसवाडे, रा. हुपरी'],
 ['सहधारक','श्री. देवाप्पा व श्री. आदिनाथ आण्णाप्पा वसवाडे'],
 ['मोजणी करणार','रमा. प्र. पाटील, उप अधीक्षक भूमि अभिलेख, हातकणंगले'],[],
 ['टिपा'],
 ['1','क्षेत्रफळ नकाशातील अधिकृत "क्षेत्राचा तपशील" तक्त्याप्रमाणे: 164.06+62.30+62.30+10.54 = 299.20 चौ.मी. = 3,220.56 चौ.फूट'],
 ['2','बाजूंची मापे मूळ नकाशावर लिहिलेली नाहीत; ती 1:500 प्रमाणानुसार हद्दरेषांच्या भूमितीवरून काढली आहेत (अचूकता ± 0.5 फूट / 0.15 मी.).'],
 ['3','या मापांवरून पुन्हा मोजलेले क्षेत्र 298.55 चौ.मी. — अधिकृत 299.20 पेक्षा 0.22% फरक.'],
 ['4','2457/4 (113.45 चौ.फूट / 10.54 चौ.मी.) तिघांच्या नावे सामाईक, पश्चिम टोकाला.'],
 ['5','पूर्वेस लागू रस्ता — 2457/1 ला 18.96 फूट, 2457/2 ला 7.74 फूट, 2457/3 ला 6.89 फूट तोंड.'],
 ['6','बांधकाम / खरेदी-विक्रीपूर्वी प्रत्यक्ष जागेवर टेप किंवा DGPS ने मापे तपासून घ्यावीत.'],
 ['7','रूपांतरण: 1 मी. = 3.28084 फूट, 1 चौ.मी. = 10.7639 चौ.फूट, 1 गुंठा = 101.17 चौ.मी.']]:
    ws4.append(r)
ws4['A1'].font=Font(bold=True,size=13); ws4['A14'].font=Font(bold=True,size=12)
ws4.column_dimensions['A'].width=20; ws4.column_dimensions['B'].width=110
for row in ws4.iter_rows():
    for c in row: c.alignment=Alignment(vertical='top',wrap_text=True)
wb.save('सि.स.नं.-2457-हुपरी-मोजमापे.xlsx')
print('saved')
